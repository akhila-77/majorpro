import re
from numpy import genfromtxt
from flask import Flask,render_template,request
from openpyxl import load_workbook
import pandas as pd
import numpy as np
from pyparsing import col
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.metrics.pairwise import cosine_similarity
import hashlib
def recommend_course(index,datasetname,num_of_rec=5):
    data=pd.read_csv(datasetname,encoding='ISO-8859-1')
    count_vect = CountVectorizer()
    cv_mat = count_vect.fit_transform(data['cname'])
    cosine_sim_mat = cosine_similarity(cv_mat)
    res=pd.DataFrame(data=data['cid'])
    idx =res.index[res['cid'] == index].tolist()[0]
    scores = list(enumerate(cosine_sim_mat[idx]))
    sorted_scores = sorted(scores,key=lambda x:x[1],reverse=True)
    selected_course_indices = [i[0] for i in sorted_scores[0:]]
    selected_course_scores = [i[1] for i in sorted_scores[0:]]
    recommended_result = data['cname'].iloc[selected_course_indices]
    rec_df = pd.DataFrame(recommended_result)
    rec_df['similarity_scores'] = selected_course_scores
    if(datasetname=='course.csv' or datasetname=='laptops.csv'):
        result1 = data['subject'].iloc[selected_course_indices]
        result2 = data['budget'].iloc[selected_course_indices]
        result3 = data['website'].iloc[selected_course_indices]
        result4=data['description'].iloc[selected_course_indices]
        ress=pd.concat([result1,result2,result3,result4],axis=1)
    elif(datasetname=='Placements.csv'):
        result1 = data['cname'].iloc[selected_course_indices]
        result2 = data['Company'].iloc[selected_course_indices]
        result3 = data['salary'].iloc[selected_course_indices]
        result4=data['Experience'].iloc[selected_course_indices]
        result5=data['ApplyLink'].iloc[selected_course_indices]
        result6=data['description'].iloc[selected_course_indices]
        ress=pd.concat([result1,result2,result3,result4,result5,result6],axis=1)
    rec_df = pd.DataFrame(ress)
    return rec_df.head(num_of_rec)
udata=pd.read_csv('unique_courses.csv')
udata=udata.to_numpy()
column_names=[]
app=Flask(__name__)
@app.route('/')
def home():
    return render_template('home.html')
@app.route('/signup',methods=['GET','POST'])
def signup():
    if request.method == "POST":
        name1=request.form['name']
        password=request.form['password']
        email=request.form['email']
        password1=(hashlib.md5(password.encode())).hexdigest()
        new_row=[name1,password1,email]
        wb = load_workbook('signup.xlsx')
        page=wb.active
        for i in range(1, page.max_row+1):
            cell_obj = page.cell(row=i, column=3)
            if(email==cell_obj.value):
                return render_template("signup.html",msg="User already registerd")
        else:
            page.append(new_row)
            wb.save(filename='signup.xlsx')
            return render_template('signup.html',msg="User Sucesfully registered")
    return render_template('signup.html')
@app.route('/login',methods=['GET','POST'])
def login():
    if request.method == "POST":
        name=request.form['name']
        pwd=request.form['password']
        password=(hashlib.md5(pwd.encode())).hexdigest()
        wb = load_workbook('signup.xlsx')
        page=wb.active
        for i in range(1, page.max_row+1):
            cell_obj = page.cell(row=i,column=1)
            passwordcell=page.cell(row=i,column=2)
            if(cell_obj.value=='admin' and password==passwordcell.value):
                return render_template('adminhome.html')
            elif(name==cell_obj.value and password==passwordcell.value):
                return render_template('index.html',name=name)
            elif(name==cell_obj.value and password!=passwordcell.value):
                return render_template('login.html',msg="Invalid password")
        else:
            return render_template('login.html',msg="User Not Registerd")
    return render_template('login.html')
@app.route('/index')
def index():
    return render_template('index.html')
@app.route('/career',methods=['GET','POST'])
def career():
    if(request.method=='POST'):
        index=request.form['submit']
        # print(int(index))
        res=recommend_course(int(index),'course.csv')
        print(res)
        res=res.to_numpy()
        column_names=['NAME','BUDGET','WEBSITE','DESCRIPTION']
        return render_template('display.html',r=res,udata=udata[0:7],colnames=column_names)
    return render_template('career.html',udata=udata[0:7])
@app.route('/laptop',methods=['GET','POST'])
def laptop():
    if(request.method=='POST'):
        index=request.form['submit']
        res=recommend_course(int(index),'laptops.csv')
        res=res.to_numpy()
        column_names=['NAME','BUDGET','WEBSITE','DESCRIPTION']
        return render_template('display.html',r=res,udata=udata[8:16],colnames=column_names)
    return render_template('laptop.html',udata=udata[8:16])
@app.route('/feedback',methods=['GET','POST'])
def feedback():
    wb = load_workbook('feedback.xlsx')
    page=wb.active
    data=[]
    for i in range(2, page.max_row+1):
        name= page.cell(row=i,column=1)
        email=page.cell(row=i,column=2)
        feedback=page.cell(row=i,column=3)
        data.append([name.value,email.value,feedback.value])
    if request.method == "POST":
        name=request.form['name']
        email=request.form['email']
        feedback=request.form['feedback']
        new_row=[name,email,feedback]
        data.append(new_row)
        page.append(new_row)
        wb.save(filename='feedback.xlsx')
        return render_template('feedback.html',msg="feedback form submitted successfully",data=data)
    return render_template('feedback.html',data=data)

@app.route('/placement',methods=['GET','POST'])
def placement():
    if(request.method=='POST'):
        index=request.form['submit']
        res=recommend_course(int(index),'Placements.csv')
        res=res.to_numpy()
        column_names=['NAME','COMPANY','SALARY','EXPERIENCE','APPLY LINK','DESCRIPTION']
        return render_template('display.html',r=res,udata=udata[15:27],colnames=column_names)
    return render_template('placement.html',udata=udata[15:27])
@app.route('/adminhome',methods=['GET','POST'])
def adminhome():
    if(request.method=='POST'):
        index=request.form['submit']
        if(index=="1"):
            data=pd.read_csv('course.csv')
            my_data = data.to_numpy()
            return render_template('admintable.html',d=my_data,value=index)
    return render_template('adminhome.html')
@app.route('/admintable',methods=['GET','POST'])
def admintable():
    if(request.method=='POST'):
        index=request.form['submit']
        return render_template('addvalue.html',value=index)
    return render_template('admintable.html')
@app.route('/addvalue')
def addvalue():
    return render_template('addvalue.html')
if __name__=='__main__':
    app.run(debug=True)